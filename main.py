import os
import re
import json
import logging
import time
import calendar
import requests
from collections import defaultdict
from io import BytesIO
from datetime import datetime

import openpyxl
import pdfplumber
import gspread
from telegram import Update
from telegram.ext import Application, MessageHandler, ContextTypes, filters
from google.oauth2.service_account import Credentials

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

BOT_TOKEN = os.getenv("BOT_TOKEN")
SPREADSHEET_ID = os.getenv("SPREADSHEET_ID")
SHEET_NAME = os.getenv("SHEET_NAME", "Реестр26")
SPREADSHEET_URL = f"https://docs.google.com/spreadsheets/d/{os.getenv('SPREADSHEET_ID')}"
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
WEBHOOK_URL = os.getenv("WEBHOOK_URL", "https://your-service.onrender.com")
PORT = int(os.getenv("PORT", "10000"))

IBAN_MAP = {
    "KZ258562203129845083": "БЦК Камкорлык",
    "KZ53722S000021350043": "Каспи Айс",
    "KZ18722S000017283463": "Каспи Багдат",
    "KZ15722C000023657799": "Каспи Голд Айко",
}

MAIN_CASH_ACCOUNTS = [
    "БЦК Камкорлык",
    "Каспи Айс",
    "Каспи Багдат",
    "Каспи Голд Айко",
    "Основная касса (Сейф)",
]

EXCLUDED_ACCOUNTS = set()

# ============ GOOGLE SHEETS ============

def get_spreadsheet():
    creds_dict = json.loads(os.getenv("GOOGLE_CREDENTIALS"))
    creds = Credentials.from_service_account_info(
        creds_dict,
        scopes=["https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive"],
    )
    client = gspread.authorize(creds)
    return client.open_by_key(SPREADSHEET_ID)


def get_sheet():
    return get_spreadsheet().worksheet(SHEET_NAME)


# ============ УТИЛИТЫ ============

def format_date(val):
    if isinstance(val, datetime):
        return val.strftime("%m/%d/%Y")
    s = str(val).replace("\n", "").strip()
    s = re.sub(r'\s+\d{1,2}:\d{2}(:\d{2})?$', '', s).strip()
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        part1, part2, y = int(m.group(1)), int(m.group(2)), m.group(3)
        if part1 > 12:
            return f"{part2:02d}/{part1:02d}/{y}"
        if part2 > 12:
            return f"{part1:02d}/{part2:02d}/{y}"
        return f"{part1:02d}/{part2:02d}/{y}"
    m = re.search(r"(\d{1,2})[.\-](\d{1,2})[.\-](\d{2,4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), m.group(3)
        if len(y) == 2:
            y = "20" + y
        return f"{mo:02d}/{d:02d}/{y}"
    return s


def get_year_from_date(date_str):
    try:
        return date_str.split("/")[2]
    except:
        return ""


def get_month_from_date(date_str):
    try:
        return int(date_str.split("/")[0])
    except:
        return ""


def get_month_nachislenia(desc, date_str):
    if desc:
        m = re.search(r"\d{1,2}[./]\d{1,2}[./]\d{2,4}", str(desc))
        if m:
            parts = re.split(r"[./]", m.group())
            if len(parts) >= 2:
                try:
                    return int(parts[1])
                except:
                    pass
        months = {"январ": 1, "феврал": 2, "март": 3, "апрел": 4, "май": 5, "мая": 5,
                  "июн": 6, "июл": 7, "август": 8, "сентябр": 9, "октябр": 10,
                  "ноябр": 11, "декабр": 12}
        for w, n in months.items():
            if w in str(desc).lower():
                return n
    try:
        return int(date_str.split("/")[0])
    except:
        return ""


def get_week(date_str):
    try:
        return datetime.strptime(date_str, "%m/%d/%Y").isocalendar()[1]
    except:
        return ""


def get_article(desc, amount):
    if not desc:
        return ""
    d = str(desc)
    for kw in ["Оплата за услуги операций по картам Kaspi Gold",
               "Оплата рекламных услуг", "Оплата за услуги процессинга Без НДС",
               "Оплата услуги по обработке данных",
               "Оплата за информационно-технологические услуги",
               "Бонусы за отзыв клиенту",
               "Оплата услуг по обработке данных, связанных с доставкой",
               "Погашение комиссии за ведение счета",
               "Бонусы клиенту от продаж за"]:
        if kw.lower() in d.lower():
            return "Комиссия за эквайринг"
    if "Расчеты РЕСПУБЛИКИ по карточкам за" in d:
        return "Оплата от покупателя, выручка"
    if "Возврат продаж с Kaspi.kz" in d:
        return "Возврат от покупателя"
    if "Возврат" in d and amount < 0:
        return "Возврат от покупателя"
    if "Возврат" in d and amount > 0:
        return "Оплата от покупателя, выручка"
    if "Продажи с Kaspi.kz" in d:
        return "Оплата от покупателя, выручка"
    return ""


def fmt_amount(val):
    try:
        f = float(val)
        return str(int(f)) if f == int(f) else str(round(f, 2))
    except:
        return str(val)


def cell_val(cell):
    return cell.value if cell and cell.value is not None else ""


def parse_num(s):
    s = str(s or "").replace(" ", "").replace("\xa0", "").replace(",", ".").replace("\n", "")
    try:
        return float(s)
    except:
        return 0


def parse_справка_num(s):
    s = str(s or "").strip().replace("\xa0", "").replace(" ", "")
    negative = s.startswith("-")
    if negative:
        s = s[1:]
    comma_count = s.count(",")
    dot_count = s.count(".")
    if comma_count > 1:
        s = s.replace(",", "")
    elif comma_count == 1 and dot_count == 0:
        parts = s.split(",")
        if len(parts[1]) == 3 and len(parts[0]) <= 3:
            s = s.replace(",", "")
        else:
            s = s.replace(",", ".")
    elif dot_count > 1:
        s = s.replace(".", "")
    try:
        result = float(s)
        return -result if negative else result
    except:
        return None


def parse_amount_from_registry(s):
    s = str(s or "").strip().replace(" ", "").replace("\xa0", "")
    if not s:
        return None
    negative = s.startswith("-")
    if negative:
        s = s[1:]
    dot_count = s.count(".")
    comma_count = s.count(",")
    if dot_count >= 1 and comma_count >= 1:
        s = s.replace(",", "")
    elif comma_count > 1:
        s = s.replace(",", "")
    elif comma_count == 1 and dot_count == 0:
        parts = s.split(",")
        if len(parts) == 2 and len(parts[1]) == 3:
            s = s.replace(",", "")
        else:
            s = s.replace(",", ".")
    elif dot_count > 1:
        s = s.replace(".", "")
    try:
        result = float(s)
        return -result if negative else result
    except:
        return None


def make_row(date_str, amount, account, desc, supplier=""):
    year = get_year_from_date(date_str)
    month_oplaty = get_month_from_date(date_str)
    week = get_week(date_str)
    month_nachislenia = get_month_nachislenia(desc, date_str)
    return [
        year,
        str(month_oplaty),
        str(week),
        date_str,
        fmt_amount(amount),
        str(month_nachislenia),
        account,
        get_article(desc, amount),
        desc,
        "",
        str(supplier) if supplier else "",
    ]


def get_last_operation_date(реестр_data) -> str:
    last_date = None
    for row in реестр_data[1:]:
        date_val = str(row[3]).strip() if len(row) > 3 else ""
        if not date_val:
            continue
        for fmt in ("%m/%d/%Y", "%d.%m.%Y", "%m/%d/%y"):
            try:
                d = datetime.strptime(date_val, fmt)
                if last_date is None or d > last_date:
                    last_date = d
                break
            except:
                pass
    return last_date.strftime("%d.%m.%Y") if last_date else datetime.now().strftime("%d.%m.%Y")


def get_last_operation_datetime(реестр_data) -> datetime:
    last_date = None
    for row in реестр_data[1:]:
        date_val = str(row[3]).strip() if len(row) > 3 else ""
        if not date_val:
            continue
        for fmt in ("%m/%d/%Y", "%d.%m.%Y", "%m/%d/%y"):
            try:
                d = datetime.strptime(date_val, fmt)
                if last_date is None or d > last_date:
                    last_date = d
                break
            except:
                pass
    return last_date if last_date else datetime.now()


# ============ ДЕДУПЛИКАЦИЯ ============

def _normalize_date(date_str: str) -> str:
    s = str(date_str).strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).strftime("%m/%d/%Y")
        except:
            pass
    parts = s.split("/")
    if len(parts) == 3:
        try:
            return f"{int(parts[0]):02d}/{int(parts[1]):02d}/{parts[2]}"
        except:
            pass
    return s


def _strip_doc_num(desc: str) -> str:
    return re.sub(r'^\d{7,12}\s+', '', str(desc).strip())


def _normalize_desc(desc: str) -> str:
    d = _strip_doc_num(desc)
    d = re.sub(r'\s+', ' ', d.replace("\n", " ").replace("\r", " ")).strip().lower()
    return d


def _normalize_amount(amount) -> str:
    amt_str = str(amount).strip().replace("\xa0", "").replace(" ", "").replace(",", "")
    try:
        return str(int(round(float(amt_str))))
    except:
        return amt_str


def _normalize_amount_exact(amount) -> str:
    amt_str = str(amount).strip().replace("\xa0", "").replace(" ", "").replace(",", "")
    try:
        return f"{float(amt_str):.2f}"
    except:
        return amt_str


def _extract_doc_num(desc_clean: str) -> str:
    desc_lower = desc_clean.lower()
    m = re.search(r'00ubs(\d+)', desc_lower)
    if m:
        return "ubs-" + m.group(1)
    m = re.match(r'^(g2-\d+)', desc_lower)
    if m:
        return m.group(1)
    m = re.match(r'^(nt-?\d+)', desc_lower)
    if m:
        return m.group(1)
    m = re.search(r'референс\s+(\d{8,12})', desc_lower)
    if m:
        return "ref-" + m.group(1)
    m = re.match(r'^(\d{7,12})\b', desc_clean)
    if m:
        return m.group(1)
    m = re.search(r'(?<!\d)(\d{7,12})(?!\d)', desc_clean)
    if m:
        return m.group(1)
    return ""


def _build_all_keys(date, amount, account, desc):
    nd = _normalize_date(str(date).strip())
    acc = str(account).strip()
    amt = _normalize_amount(amount)
    amt_exact = _normalize_amount_exact(amount)
    desc_raw = re.sub(r'\s+', ' ', str(desc).replace("\n", " ").replace("\r", " ")).strip()
    desc_norm = _normalize_desc(desc_raw)
    desc_short = desc_norm[:50]
    doc_num = _extract_doc_num(desc_raw)
    keys = []
    if doc_num:
        keys.append(("doc", nd, acc, doc_num))
    keys.append(("exact", nd, acc, amt_exact))
    keys.append(("amt", nd, acc, amt))
    if desc_short:
        keys.append(("text", nd, acc, desc_short))
    if desc_short:
        keys.append(("amt+text", nd, acc, amt, desc_short))
    return keys


def is_duplicate(r, existing_keys: set) -> bool:
    keys = _build_all_keys(r[3], r[4], r[6], r[8])
    for k in keys:
        if k in existing_keys:
            return True
    return False


def build_existing_keys(existing_data) -> tuple:
    key_to_row = {}
    for i, row in enumerate(existing_data[1:], start=2):
        if len(row) < 9:
            continue
        if not any(str(c).strip() for c in row):
            continue
        date_val = str(row[3]).strip()
        amount_val = str(row[4]).strip().replace(",", "").replace(" ", "").replace("\xa0", "")
        account_val = str(row[6]).strip()
        desc_val = str(row[8]).strip()
        if not date_val or not amount_val or not account_val:
            continue
        keys = _build_all_keys(date_val, amount_val, account_val, desc_val)
        for k in keys:
            if k not in key_to_row:
                key_to_row[k] = i
    return key_to_row, set(key_to_row.keys())


def find_existing_row(r, key_to_row: dict):
    keys = _build_all_keys(r[3], r[4], r[6], r[8])
    for k in keys:
        if k in key_to_row:
            return key_to_row[k]
    return None


def append_rows_from_col_a(sheet, rows):
    all_vals = sheet.get_all_values()
    last_row = 1
    for i, row in enumerate(all_vals, start=1):
        col_a = str(row[0]).strip() if len(row) > 0 else ""
        col_d = str(row[3]).strip() if len(row) > 3 else ""
        has_year = bool(re.match(r'^\d{4}$', col_a))
        has_date = bool(re.match(r'^\d{2}/\d{2}/\d{4}$', col_d))
        if has_year or has_date:
            last_row = i
    start_row = last_row + 1
    if not rows:
        return start_row
    end_col = "K"
    range_name = f"A{start_row}:{end_col}{start_row + len(rows) - 1}"
    sheet.update(range_name=range_name, values=rows, value_input_option="USER_ENTERED")
    return start_row


# ============ СВЕРКА ОСТАТКОВ ============

ACCOUNT_SYNONYMS = {
    "халык":       ["халык", "народный", "halyk", "hsbk"],
    "каспи":       ["каспи", "kaspi", "каспий"],
    "бцк":         ["бцк", "центркредит", "bcc", "kcjb"],
    "серик":       ["серик", "серік"],
    "имангазиева": ["имангазиева"],
    "орынбаева":   ["орынбаева"],
    "дильназ":     ["дильназ"],
    "коко":        ["коко", "сулейменов"],
    "арман":       ["арман"],
    "голд":        ["голд", "gold"],
    "тоо":         ["тоо"],
    "ип":          ["ип", "ip"],
    "депозит":     ["депозит", "deposit"],
    "usd":         ["usd", "доллар"],
    "касса":       ["касса", "сейф", "наличные", "cash", "кассa"],
    "камкорлык":   ["камкорлык"],
    "айс":         ["айс"],
    "багдат":      ["багдат"],
    "айко":        ["айко"],
}

BANK_CONFLICT_GROUPS = [
    {"халык", "народный"},
    {"каспи"},
    {"бцк", "центркредит"},
    {"втб"},
]


def _normalize_tokens(name: str) -> set:
    name = name.lower().strip()
    name = re.sub(r'[«»"\'(),.\-]', " ", name)
    words = name.split()
    tokens = set()
    for word in words:
        tokens.add(word)
        for canon, synonyms in ACCOUNT_SYNONYMS.items():
            if word in synonyms:
                tokens.add(canon)
                break
    return tokens


def _account_similarity(name_a: str, name_b: str) -> float:
    ta = _normalize_tokens(name_a)
    tb = _normalize_tokens(name_b)
    if not ta or not tb:
        return 0.0
    score = len(ta & tb) / max(len(ta), len(tb))
    for group in BANK_CONFLICT_GROUPS:
        a_in = bool(ta & group)
        b_in = bool(tb & group)
        a_other = any(ta & g for g in BANK_CONFLICT_GROUPS if g != group)
        b_other = any(tb & g for g in BANK_CONFLICT_GROUPS if g != group)
        if a_other and b_in and not (ta & group):
            score *= 0.3
        if b_other and a_in and not (tb & group):
            score *= 0.3
    return score


def find_account_in_справка(account_name: str, справка_data: list):
    search = account_name.strip().lower()
    for row in справка_data:
        if not row or not row[0].strip():
            continue
        candidate = row[0].strip()
        if candidate.lower() == search:
            balance = parse_справка_num(row[1] if len(row) > 1 else "")
            if balance is not None:
                return candidate, balance
    best_name = None
    best_balance = None
    best_score = 0.0
    for row in справка_data:
        if not row or not row[0].strip():
            continue
        candidate = row[0].strip()
        score = _account_similarity(account_name, candidate)
        if score > best_score:
            best_score = score
            best_name = candidate
            best_balance = parse_справка_num(row[1] if len(row) > 1 else "")
    if best_score >= 0.35 and best_balance is not None:
        logger.info(f"find_account_in_справка: '{account_name}' -> '{best_name}' score={best_score:.2f}")
        return best_name, best_balance
    logger.warning(f"find_account_in_справка: '{account_name}' не найден (best={best_score:.2f})")
    return None, None


def _clean_name_for_match(s):
    import unicodedata
    s = str(s)
    s = s.replace("\xa0", " ").replace("\u200b", "").replace("\n", " ").replace("\r", " ")
    s = unicodedata.normalize("NFKC", s)
    s = s.strip("'\"")
    return " ".join(s.split()).lower()


def _matches_account_strict(row_acc, target):
    return _clean_name_for_match(row_acc) == _clean_name_for_match(target)


def _matches_account(row_acc, target):
    r = _clean_name_for_match(row_acc)
    t = _clean_name_for_match(target)
    if r == t:
        return True
    if _account_similarity(r, t) >= 0.80:
        return True
    return False


def _load_initial_balances():
    try:
        справка = get_spreadsheet().worksheet("Счета2026(Справка)")
        справка_data = справка.get_all_values()
        result = {}
        for row in справка_data:
            if row and row[0].strip() and len(row) > 1:
                bal = parse_справка_num(row[1])
                if bal is not None:
                    result[row[0].strip()] = bal
        return result
    except Exception as e:
        logger.warning(f"_load_initial_balances error: {e}")
        return {}


def _get_initial_for_account(account_name, initial_balances):
    for name, bal in initial_balances.items():
        if name.lower() == account_name.lower():
            return bal
    best_score = 0.0
    best_bal = 0.0
    for name, bal in initial_balances.items():
        score = _account_similarity(account_name, name)
        if score > best_score and score >= 0.35:
            best_score = score
            best_bal = bal
    return best_bal


def get_account_balance(account_name: str):
    spreadsheet = get_spreadsheet()
    initial_balances = _load_initial_balances()
    initial = _get_initial_for_account(account_name, initial_balances)
    реестр = spreadsheet.worksheet(SHEET_NAME)
    реестр_data = реестр.get_all_values()
    ops_total = 0.0
    ops_count = 0
    for row in реестр_data[1:]:
        acc_val = str(row[6]).strip() if len(row) > 6 else ""
        amt_val = str(row[4]).strip() if len(row) > 4 else ""
        if not acc_val or not amt_val:
            continue
        if _matches_account_strict(acc_val, account_name):
            parsed = parse_amount_from_registry(amt_val)
            if parsed is not None:
                ops_total += parsed
                ops_count += 1
    dds = round(initial + ops_total, 2)
    return initial, round(ops_total, 2), dds, ops_count


def get_account_balance_on_date(account_name: str, target_date_str: str):
    target_date = None
    for fmt in ("%d.%m.%Y", "%m/%d/%Y", "%d.%m.%y", "%Y-%m-%d"):
        try:
            target_date = datetime.strptime(target_date_str.strip(), fmt)
            break
        except:
            pass
    if not target_date:
        return None, None, None, f"Не удалось распознать дату: {target_date_str}"
    initial_balances = _load_initial_balances()
    initial = _get_initial_for_account(account_name, initial_balances)
    реестр = get_spreadsheet().worksheet(SHEET_NAME)
    реестр_data = реестр.get_all_values()
    ops_total = 0.0
    ops_count = 0
    for row in реестр_data[1:]:
        acc_val = str(row[6]).strip() if len(row) > 6 else ""
        amt_val = str(row[4]).strip() if len(row) > 4 else ""
        date_val = str(row[3]).strip() if len(row) > 3 else ""
        if not acc_val or not amt_val or not date_val:
            continue
        if not _matches_account_strict(acc_val, account_name):
            continue
        row_date = None
        for fmt in ("%m/%d/%Y", "%d.%m.%Y", "%m/%d/%y"):
            try:
                row_date = datetime.strptime(date_val, fmt)
                break
            except:
                pass
        if row_date and row_date <= target_date:
            parsed = parse_amount_from_registry(amt_val)
            if parsed is not None:
                ops_total += parsed
                ops_count += 1
    return initial, round(ops_total, 2), round(initial + ops_total, 2), ops_count


def find_operation_by_amount(target_amount: float, tolerance: float = 1.0):
    реестр = get_spreadsheet().worksheet(SHEET_NAME)
    реестр_data = реестр.get_all_values()
    matches = []
    for row in реестр_data[1:]:
        acc_val  = str(row[6]).strip() if len(row) > 6 else ""
        amt_val  = str(row[4]).strip() if len(row) > 4 else ""
        date_val = str(row[3]).strip() if len(row) > 3 else ""
        desc_val = str(row[8]).strip() if len(row) > 8 else ""
        if not acc_val or not amt_val or not date_val:
            continue
        parsed = parse_amount_from_registry(amt_val)
        if parsed is None:
            continue
        if abs(abs(parsed) - abs(target_amount)) <= tolerance:
            row_date = None
            for fmt in ("%m/%d/%Y", "%d.%m.%Y", "%m/%d/%y"):
                try:
                    row_date = datetime.strptime(date_val, fmt)
                    break
                except:
                    pass
            date_display = row_date.strftime("%d.%m.%Y") if row_date else date_val
            matches.append((acc_val, date_display, parsed, desc_val[:60]))

    def sort_key(m):
        try:
            return datetime.strptime(m[1], "%d.%m.%Y")
        except:
            return datetime.min
    matches.sort(key=sort_key)
    return matches


def find_date_by_balance(target_amount: float, tolerance: float = 1.0):
    initial_balances = _load_initial_balances()
    реестр = get_spreadsheet().worksheet(SHEET_NAME)
    реестр_data = реестр.get_all_values()

    ops_by_account = defaultdict(list)
    for row in реестр_data[1:]:
        acc_val  = str(row[6]).strip() if len(row) > 6 else ""
        amt_val  = str(row[4]).strip() if len(row) > 4 else ""
        date_val = str(row[3]).strip() if len(row) > 3 else ""
        if not acc_val or not amt_val or not date_val:
            continue
        row_date = None
        for fmt in ("%m/%d/%Y", "%d.%m.%Y", "%m/%d/%y"):
            try:
                row_date = datetime.strptime(date_val, fmt)
                break
            except:
                pass
        if not row_date:
            continue
        parsed = parse_amount_from_registry(amt_val)
        if parsed is not None:
            ops_by_account[acc_val].append((row_date, parsed))

    matches = []
    for account, ops in ops_by_account.items():
        initial = _get_initial_for_account(account, initial_balances)
        ops_sorted = sorted(ops, key=lambda x: x[0])
        unique_dates = sorted(set(d for d, _ in ops_sorted))
        for check_date in unique_dates:
            total = initial + sum(amt for dt, amt in ops_sorted if dt <= check_date)
            balance = round(total, 2)
            if abs(balance - target_amount) <= tolerance:
                matches.append((account, check_date.strftime("%d.%m.%Y"), balance))
    return matches


def find_date_by_daily_total(target_amount: float, tolerance: float = 1.0):
    реестр = get_spreadsheet().worksheet(SHEET_NAME)
    реестр_data = реестр.get_all_values()

    all_ops = []
    for row in реестр_data[1:]:
        acc_val  = str(row[6]).strip() if len(row) > 6 else ""
        amt_val  = str(row[4]).strip() if len(row) > 4 else ""
        date_val = str(row[3]).strip() if len(row) > 3 else ""
        if not acc_val or not amt_val or not date_val:
            continue
        row_date = None
        for fmt in ("%m/%d/%Y", "%d.%m.%Y", "%m/%d/%y"):
            try:
                row_date = datetime.strptime(date_val, fmt)
                break
            except:
                pass
        if not row_date:
            continue
        parsed = parse_amount_from_registry(amt_val)
        if parsed is not None:
            all_ops.append((row_date, acc_val, parsed))

    if not all_ops:
        return []

    matches = []

    all_ops_sorted = sorted(all_ops, key=lambda x: x[0])
    unique_dates = sorted(set(d for d, _, _ in all_ops_sorted))
    cumulative = 0.0
    for check_date in unique_dates:
        day_sum = sum(amt for dt, _, amt in all_ops_sorted if dt == check_date)
        cumulative = round(cumulative + day_sum, 2)
        if abs(cumulative - target_amount) <= tolerance:
            matches.append(("Все счета (нарастающий)", check_date.strftime("%d.%m.%Y"), cumulative))

    ops_by_account = defaultdict(list)
    for dt, acc, amt in all_ops:
        ops_by_account[acc].append((dt, amt))

    for account, ops in ops_by_account.items():
        ops_sorted = sorted(ops, key=lambda x: x[0])
        unique_acc_dates = sorted(set(d for d, _ in ops_sorted))
        cumulative_acc = 0.0
        for check_date in unique_acc_dates:
            day_sum = sum(amt for dt, amt in ops_sorted if dt == check_date)
            cumulative_acc = round(cumulative_acc + day_sum, 2)
            if abs(cumulative_acc - target_amount) <= tolerance:
                matches.append((account, check_date.strftime("%d.%m.%Y"), cumulative_acc))

    daily_totals = defaultdict(float)
    for dt, _, amt in all_ops:
        daily_totals[dt] = round(daily_totals[dt] + amt, 2)
    for check_date, total in daily_totals.items():
        if abs(total - target_amount) <= tolerance:
            matches.append(("Дневной оборот (все счета)", check_date.strftime("%d.%m.%Y"), total))

    daily_by_acc = defaultdict(lambda: defaultdict(float))
    for dt, acc, amt in all_ops:
        daily_by_acc[acc][dt] = round(daily_by_acc[acc][dt] + amt, 2)
    for account, days in daily_by_acc.items():
        for check_date, total in days.items():
            if abs(total - target_amount) <= tolerance:
                matches.append((f"{account} (дневной оборот)", check_date.strftime("%d.%m.%Y"), total))

    def sort_key(m):
        try:
            return datetime.strptime(m[1], "%d.%m.%Y")
        except:
            return datetime.min
    matches.sort(key=sort_key)
    return matches


def get_main_cash_summary():
    initial_balances = _load_initial_balances()
    реестр = get_spreadsheet().worksheet(SHEET_NAME)
    реестр_data = реестр.get_all_values()

    data_date = get_last_operation_date(реестр_data)

    ops_by_account = {acc: 0.0 for acc in MAIN_CASH_ACCOUNTS}
    count_by_account = {acc: 0 for acc in MAIN_CASH_ACCOUNTS}

    for row in реестр_data[1:]:
        acc_val = str(row[6]).strip() if len(row) > 6 else ""
        amt_val = str(row[4]).strip() if len(row) > 4 else ""
        if not acc_val or not amt_val:
            continue
        if acc_val in EXCLUDED_ACCOUNTS:
            continue
        for target in MAIN_CASH_ACCOUNTS:
            if _matches_account_strict(acc_val, target):
                parsed = parse_amount_from_registry(amt_val)
                if parsed is not None:
                    ops_by_account[target] += parsed
                    count_by_account[target] += 1
                break

    lines = []
    total = 0.0
    for acc in MAIN_CASH_ACCOUNTS:
        initial = _get_initial_for_account(acc, initial_balances)
        ops = ops_by_account[acc]
        current = round(initial + ops, 2)
        total += current
        lines.append(
            f"  {acc}: {current:,.0f} тг"
            f"  (нач.: {initial:,.0f} + обороты: {ops:,.0f}, {count_by_account[acc]} оп.)"
        )

    result = f"Остатки по всем счетам на {data_date}:\n"
    result += "\n".join(lines)
    result += f"\n{'─' * 40}\n"
    result += f"ИТОГО ПО ВСЕМ СЧЕТАМ: {total:,.0f} тг"
    return result, round(total, 2)


def _last_day_of_month(year: int, month: int) -> datetime:
    last_day = calendar.monthrange(year, month)[1]
    return datetime(year, month, last_day, 23, 59, 59)


def _resolve_target_date_for_month(month: int, year: int) -> tuple:
    now = datetime.now()
    current_year = now.year
    current_month = now.month

    if (year < current_year) or (year == current_year and month < current_month):
        target_dt = _last_day_of_month(year, month)
        label = target_dt.strftime("%d.%m.%Y")
        return target_dt, label

    реестр = get_spreadsheet().worksheet(SHEET_NAME)
    реестр_data = реестр.get_all_values()
    last_op_dt = get_last_operation_datetime(реестр_data)
    label = last_op_dt.strftime("%d.%m.%Y") + " (последняя операция)"
    return last_op_dt, label


def get_main_cash_summary_on_date(target_date: datetime):
    initial_balances = _load_initial_balances()
    реестр = get_spreadsheet().worksheet(SHEET_NAME)
    реестр_data = реестр.get_all_values()

    ops_by_account = {acc: 0.0 for acc in MAIN_CASH_ACCOUNTS}
    count_by_account = {acc: 0 for acc in MAIN_CASH_ACCOUNTS}

    for row in реестр_data[1:]:
        acc_val  = str(row[6]).strip() if len(row) > 6 else ""
        amt_val  = str(row[4]).strip() if len(row) > 4 else ""
        date_val = str(row[3]).strip() if len(row) > 3 else ""
        if not acc_val or not amt_val or not date_val:
            continue
        if acc_val in EXCLUDED_ACCOUNTS:
            continue
        row_date = None
        for fmt in ("%m/%d/%Y", "%d.%m.%Y", "%m/%d/%y"):
            try:
                row_date = datetime.strptime(date_val, fmt)
                break
            except:
                pass
        if not row_date or row_date > target_date:
            continue
        for target_acc in MAIN_CASH_ACCOUNTS:
            if _matches_account_strict(acc_val, target_acc):
                parsed = parse_amount_from_registry(amt_val)
                if parsed is not None:
                    ops_by_account[target_acc] += parsed
                    count_by_account[target_acc] += 1
                break

    date_label = target_date.strftime("%d.%m.%Y")
    lines = []
    total = 0.0
    for acc in MAIN_CASH_ACCOUNTS:
        initial = _get_initial_for_account(acc, initial_balances)
        ops = ops_by_account[acc]
        current = round(initial + ops, 2)
        total += current
        lines.append(
            f"  {acc}: {current:,.0f} тг"
            f"  (нач.: {initial:,.0f} + обороты: {ops:,.0f}, {count_by_account[acc]} оп.)"
        )

    result = f"Остатки по всем счетам на {date_label}:\n"
    result += "\n".join(lines)
    result += f"\n{'─' * 40}\n"
    result += f"ИТОГО ПО ВСЕМ СЧЕТАМ: {total:,.0f} тг"
    return result, round(total, 2)


def build_balance_msg(account, bank_closing_balance):
    if bank_closing_balance is None:
        return "\n⚠️ Исходящий остаток не найден в файле"
    try:
        initial, ops_total, dds, ops_count = get_account_balance(account)
        bank_balance = round(bank_closing_balance, 2)
        diff = round(bank_balance - dds, 2)
        msg = ""
        if abs(diff) < 1:
            msg += f"\n✅ Остаток сходится: {bank_balance:,.2f} ₸"
        else:
            msg += f"\n❌ Остаток НЕ сходится!\n"
            msg += f"  Банк: {bank_balance:,.2f} ₸\n"
            msg += f"  ДДС:  {dds:,.2f} ₸\n"
            msg += f"  Разница: {diff:,.2f} ₸"
        msg += f"\n\n📊 Расчёт ДДС:\n"
        msg += f"  Нач. остаток (Справка): {initial:,.2f} ₸\n"
        msg += f"  + Операции ({ops_count} строк): {ops_total:,.2f} ₸\n"
        msg += f"  = Итого ДДС: {dds:,.2f} ₸\n"
        msg += f"\n🏦 Банк (исходящий): {bank_balance:,.2f} ₸"
        if abs(diff) >= 1:
            msg += f"\n\n🔎 Требуется проверка операций и дублей."
        return msg
    except Exception as e:
        return f"\n⚠️ Ошибка сверки: {e}"


def get_account_rows(account_name):
    реестр = get_spreadsheet().worksheet(SHEET_NAME)
    реестр_data = реестр.get_all_values()
    matched = []
    for i, row in enumerate(реестр_data[1:], start=2):
        acc_val = row[6].strip() if len(row) > 6 else ""
        if acc_val and _matches_account(acc_val, account_name):
            matched.append((i, row))
    return matched


# ============ ОБОРОТЫ ЗА ПЕРИОД ============

def get_turnover_for_range(start_date_str: str, end_date_str: str):
    start_date = None
    end_date = None
    for fmt in ("%d.%m.%Y", "%m/%d/%Y", "%d.%m.%y"):
        if not start_date:
            try:
                start_date = datetime.strptime(start_date_str.strip(), fmt)
            except:
                pass
        if not end_date:
            try:
                end_date = datetime.strptime(end_date_str.strip(), fmt)
            except:
                pass
    if not start_date or not end_date:
        return f"Не удалось распознать даты: {start_date_str} — {end_date_str}"

    реестр = get_spreadsheet().worksheet(SHEET_NAME)
    реестр_data = реестр.get_all_values()

    income = 0.0
    expense = 0.0
    ops_count = 0

    for row in реестр_data[1:]:
        acc_val  = str(row[6]).strip() if len(row) > 6 else ""
        amt_val  = str(row[4]).strip() if len(row) > 4 else ""
        date_val = str(row[3]).strip() if len(row) > 3 else ""
        if not acc_val or not amt_val or not date_val:
            continue
        if acc_val in EXCLUDED_ACCOUNTS:
            continue
        row_date = None
        for fmt in ("%m/%d/%Y", "%d.%m.%Y", "%m/%d/%y"):
            try:
                row_date = datetime.strptime(date_val, fmt)
                break
            except:
                pass
        if not row_date:
            continue
        if not (start_date.date() <= row_date.date() <= end_date.date()):
            continue
        parsed = parse_amount_from_registry(amt_val)
        if parsed is None:
            continue
        if parsed > 0:
            income += parsed
        else:
            expense += parsed
        ops_count += 1

    s = start_date.strftime("%d.%m.%Y")
    e = end_date.strftime("%d.%m.%Y")
    label = f"за {s}" if s == e else f"с {s} по {e}"
    return "\n".join([
        f"Обороты {label} ({ops_count} операций):",
        f"  Приход:  +{income:,.0f} тг",
        f"  Расход:  {expense:,.0f} тг",
        f"  Нетто:   {income + expense:,.0f} тг",
    ])


def get_daily_summary_for_period(start_date_str: str, end_date_str: str):
    start_date = None
    end_date = None
    for fmt in ("%d.%m.%Y", "%m/%d/%Y", "%d.%m.%y"):
        if not start_date:
            try:
                start_date = datetime.strptime(start_date_str.strip(), fmt)
            except:
                pass
        if not end_date:
            try:
                end_date = datetime.strptime(end_date_str.strip(), fmt)
            except:
                pass
    if not start_date or not end_date:
        return f"Не удалось распознать даты: {start_date_str} — {end_date_str}"

    initial_balances = _load_initial_balances()
    реестр = get_spreadsheet().worksheet(SHEET_NAME)
    реестр_data = реестр.get_all_values()

    all_ops = []
    for row in реестр_data[1:]:
        acc_val  = str(row[6]).strip() if len(row) > 6 else ""
        amt_val  = str(row[4]).strip() if len(row) > 4 else ""
        date_val = str(row[3]).strip() if len(row) > 3 else ""
        if not acc_val or not amt_val or not date_val:
            continue
        if acc_val in EXCLUDED_ACCOUNTS:
            continue
        row_date = None
        for fmt in ("%m/%d/%Y", "%d.%m.%Y", "%m/%d/%y"):
            try:
                row_date = datetime.strptime(date_val, fmt)
                break
            except:
                pass
        if not row_date or row_date > end_date:
            continue
        parsed = parse_amount_from_registry(amt_val)
        if parsed is not None:
            all_ops.append((row_date, acc_val, parsed))

    period_dates = sorted(set(
        d.date() for d, _, _ in all_ops
        if start_date.date() <= d.date() <= end_date.date()
    ))

    if not period_dates:
        return f"Операций в периоде {start_date_str} — {end_date_str} не найдено."

    lines = [f"Ежедневные итоги за {start_date.strftime('%d.%m.%Y')} — {end_date.strftime('%d.%m.%Y')}:"]
    lines.append(f"{'Дата':<12} {'Приход':>14} {'Расход':>14} {'Нетто':>14} {'Итого счета':>16}")
    lines.append("─" * 72)

    for check_date in period_dates:
        check_dt = datetime(check_date.year, check_date.month, check_date.day, 23, 59, 59)
        day_income = sum(amt for dt, _, amt in all_ops if dt.date() == check_date and amt > 0)
        day_expense = sum(amt for dt, _, amt in all_ops if dt.date() == check_date and amt < 0)
        total_balance = 0.0
        for acc in MAIN_CASH_ACCOUNTS:
            initial = _get_initial_for_account(acc, initial_balances)
            ops_sum = sum(amt for dt, a, amt in all_ops if _matches_account_strict(a, acc) and dt <= check_dt)
            total_balance += initial + ops_sum
        netto = day_income + day_expense
        lines.append(
            f"{check_date.strftime('%d.%m.%Y'):<12}"
            f" {day_income:>+14,.0f}"
            f" {day_expense:>14,.0f}"
            f" {netto:>+14,.0f}"
            f" {total_balance:>16,.0f}"
        )

    return "\n".join(lines)


# ============ ДАННЫЕ ДЛЯ ИИ ============

def get_sheets_data_for_ai(filter_account=None, filter_month=None, limit_rows=100):
    try:
        spreadsheet = get_spreadsheet()
        реестр = spreadsheet.worksheet(SHEET_NAME)
        data = реестр.get_all_values()
        initial_balances = _load_initial_balances()

        if len(data) < 2:
            return "Данных в таблице пока нет.", []

        rows = data[1:]
        data_date = get_last_operation_date(data)

        month_totals = {}
        account_totals = {}
        article_totals = {}
        filtered_month_totals = {}
        filtered_account_totals = {}
        filtered_article_totals = {}
        detail_rows = []

        month_names = {
            "1": "Январь", "2": "Февраль", "3": "Март", "4": "Апрель",
            "5": "Май", "6": "Июнь", "7": "Июль", "8": "Август",
            "9": "Сентябрь", "10": "Октябрь", "11": "Ноябрь", "12": "Декабрь"
        }
        has_filter = bool(filter_month or filter_account)

        for row in rows:
            try:
                month = row[1] if len(row) > 1 else ""
                amount_str = row[4] if len(row) > 4 else "0"
                account = row[6] if len(row) > 6 else ""
                article = row[7] if len(row) > 7 else ""
                date_str = row[3] if len(row) > 3 else ""
                desc = row[8] if len(row) > 8 else ""

                if account in EXCLUDED_ACCOUNTS:
                    continue

                amount = parse_amount_from_registry(amount_str)
                if amount is None:
                    continue
                if month:
                    month_totals[month] = month_totals.get(month, 0) + amount
                if account:
                    account_totals[account] = account_totals.get(account, 0) + amount
                if article:
                    article_totals[article] = article_totals.get(article, 0) + amount
                match = True
                if filter_account and filter_account.lower() not in account.lower():
                    match = False
                if filter_month and str(filter_month) != str(month):
                    match = False
                if match:
                    if month:
                        filtered_month_totals[month] = filtered_month_totals.get(month, 0) + amount
                    if account:
                        filtered_account_totals[account] = filtered_account_totals.get(account, 0) + amount
                    if article:
                        filtered_article_totals[article] = filtered_article_totals.get(article, 0) + amount
                    detail_rows.append({
                        "дата": date_str,
                        "месяц": month_names.get(month, month),
                        "счёт": account,
                        "сумма": amount_str,
                        "статья": article,
                        "описание": desc[:80] if desc else "",
                    })
            except:
                continue

        use_month_totals   = filtered_month_totals   if has_filter else month_totals
        use_account_totals = filtered_account_totals if has_filter else account_totals
        use_article_totals = filtered_article_totals if has_filter else article_totals

        summary = f"Данные по состоянию на {data_date}. Всего строк в реестре: {len(rows)}\n"
        if has_filter:
            parts = []
            if filter_month:
                parts.append(f"месяц={month_names.get(str(filter_month), filter_month)}")
            if filter_account:
                parts.append(f"счёт={filter_account}")
            summary += f"Применён фильтр: {', '.join(parts)}\n"
            summary += f"Строк по фильтру: {len(detail_rows)}\n"
        summary += "\n"

        summary += "ОБОРОТЫ ПО МЕСЯЦАМ:\n"
        if use_month_totals:
            for m in sorted(use_month_totals.keys(), key=lambda x: int(x) if x.isdigit() else 99):
                summary += f"  {month_names.get(m, f'Месяц {m}')}: {use_month_totals[m]:,.0f} ₸\n"
        else:
            summary += "  (нет данных)\n"

        summary += f"\nТЕКУЩИЕ ОСТАТКИ ПО КАЖДОМУ СЧЕТУ (на {data_date}):\n"
        total_all = 0.0
        for acc, ops_total in sorted(account_totals.items(), key=lambda x: x[0]):
            initial = _get_initial_for_account(acc, initial_balances)
            current = initial + ops_total
            total_all += current
            if initial != 0:
                summary += f"  {acc}: {current:,.0f} ₸  (нач.: {initial:,.0f} + обороты: {ops_total:,.0f})\n"
            else:
                summary += f"  {acc}: {current:,.0f} ₸  (нач. остаток не найден, обороты: {ops_total:,.0f})\n"
        summary += f"  ИТОГО НА ВСЕХ СЧЕТАХ: {total_all:,.0f} ₸\n"

        summary += "\nОБОРОТЫ ПО СЧЕТАМ"
        if has_filter:
            summary += " (по фильтру)"
        summary += ":\n"
        for acc, ops_total in sorted(use_account_totals.items(), key=lambda x: -abs(x[1])):
            summary += f"  {acc}: {ops_total:,.0f} ₸\n"

        summary += "\nПО СТАТЬЯМ"
        if has_filter:
            summary += " (по фильтру)"
        summary += ":\n"
        for art, total in sorted(use_article_totals.items(), key=lambda x: -abs(x[1])):
            if art:
                summary += f"  {art}: {total:,.0f} ₸\n"

        return summary, detail_rows[:limit_rows]

    except Exception as e:
        return f"Ошибка получения данных: {e}", []


# ============ ДЕТЕКТОР ВОПРОСОВ ============

LIST_ACCOUNTS_KEYWORDS = [
    "список счетов", "напиши все счета", "напиши счета", "покажи все счета",
    "покажи счета", "какие счета", "какие есть счета", "перечисли счета",
    "перечисли все счета", "все названия счетов", "названия счетов",
    "какие у вас счета", "какие счёта", "все счёта", "покажи счёта",
]

SEYF_KEYWORDS = [
    "основная касса", "основной кассе", "основную кассу", "основной кассы",
    "основной кассой", "сейф", "сейфе", "сейфа", "сейфом",
    "остаток кассы", "остаток по кассе",
]

ALL_ACCOUNTS_KEYWORDS = [
    "сколько денег", "денег всего", "всего денег", "сколько всего денег",
    "на всех счетах", "по всем счетам", "общий остаток",
    "итого по счетам", "итого на счетах",
    "сколько у нас денег", "сколько денег у нас", "сколько у нас на счетах",
    "общий баланс", "суммарный остаток", "суммарный баланс",
    "итого по всем", "общая сумма",
]

BALANCE_ON_MONTH_KEYWORDS = [
    "остаток на",
    "остатки на",
    "какой остаток",
    "баланс на",
    "сколько было на",
    "сколько денег было в",
    "остаток в",
    "остатки в",
    "остаток за",
    "остатки за",
    "сколько было в",
    "на конец",
    "в конце",
    "к концу месяца",
]

BALANCE_SEARCH_KEYWORDS = [
    "какого дня", "какой день", "какого числа", "в какой день",
    "когда был остаток", "когда был баланс", "когда была сумма", "когда стало",
    "дата остатка", "найди дату", "какая дата", "какого числа был",
    "когда на счете было", "когда на счёте было",
    "этот остаток", "этот баланс", "такой остаток",
    "когда было столько", "когда была такая сумма",
]

TURNOVER_INTENT_KEYWORDS = [
    "потрачено", "потратили", "потрачено за", "потрачено на", "потрачено к",
    "израсходовано", "расход за", "расход на", "расход с",
    "приход за", "приход на", "приход с",
    "обороты за", "обороты на", "обороты с",
    "сколько пришло", "сколько ушло",
    "сколько потрачено", "сколько израсходовано",
    "сколько денег потрачено", "сколько было потрачено",
    "сколько денег было потрачено", "сколько денег ушло",
    "сколько денег пришло", "сколько денег было потрачено",
    "операции за", "операции с", "транзакции за",
]

MONTH_KEYWORDS = {
    "январ": 1, "феврал": 2, "март": 3, "апрел": 4,
    "май": 5, "мая": 5, "июн": 6, "июл": 7, "август": 8,
    "сентябр": 9, "октябр": 10, "ноябр": 11, "декабр": 12,
}

BALANCE_AT_DATE_KEYWORDS = [
    "сколько было", "сколько денег было", "остаток на конец", "остаток в конце",
    "остатки на", "сколько на счетах было", "сколько у нас было",
    "на всех счетах было", "баланс на конец", "баланс в",
]


def _extract_month_from_question(text: str):
    t = text.lower()
    for kw, num in MONTH_KEYWORDS.items():
        if kw in t:
            return num
    m = re.search(r'\b(\d{1,2})\s*месяц', t)
    if m:
        val = int(m.group(1))
        if 1 <= val <= 12:
            return val
    return None


def _extract_dates_from_question(text: str):
    dates = []
    for m in re.finditer(r'\b(\d{1,2})[./](\d{1,2})[./](\d{2,4})\b', text):
        d, mo, y = int(m.group(1)), int(m.group(2)), m.group(3)
        if len(y) == 2:
            y = "20" + y
        dates.append(f"{d:02d}.{mo:02d}.{y}")

    day_month_map = {
        "январ": 1, "феврал": 2, "март": 3, "апрел": 4,
        "май": 5, "мая": 5, "июн": 6, "июл": 7, "август": 8,
        "сентябр": 9, "октябр": 10, "ноябр": 11, "декабр": 12,
    }
    t = text.lower()
    year = datetime.now().year
    for kw, month_num in day_month_map.items():
        for m in re.finditer(rf'\b(\d{{1,2}})\s+{kw}\w*', t):
            day = int(m.group(1))
            candidate = f"{day:02d}.{month_num:02d}.{year}"
            if candidate not in dates:
                dates.append(candidate)

    return dates[:2]


def _extract_date_from_question(text: str):
    dates = _extract_dates_from_question(text)
    return dates[0] if dates else None


def _is_balance_at_date_question(text: str) -> bool:
    t = text.lower()
    return any(kw in t for kw in BALANCE_AT_DATE_KEYWORDS)


def _is_list_accounts_question(text: str) -> bool:
    t = text.lower().strip()
    return any(kw in t for kw in LIST_ACCOUNTS_KEYWORDS)


def _is_seyf_question(text: str) -> bool:
    t = text.lower().strip()
    return any(kw in t for kw in SEYF_KEYWORDS)


def _is_all_accounts_question(text: str) -> bool:
    t = text.lower().strip()
    return any(kw in t for kw in ALL_ACCOUNTS_KEYWORDS)


def _is_balance_search_question(text: str) -> bool:
    t = text.lower().strip()
    return any(kw in t for kw in BALANCE_SEARCH_KEYWORDS)


def _is_turnover_intent_question(text: str) -> bool:
    t = text.lower().strip()
    return any(kw in t for kw in TURNOVER_INTENT_KEYWORDS)


def _is_balance_on_month_question(text: str) -> bool:
    t = text.lower().strip()
    has_month = _extract_month_from_question(t) is not None
    if not has_month:
        return False
    if _is_turnover_intent_question(t):
        return False
    return any(kw in t for kw in BALANCE_ON_MONTH_KEYWORDS)


# ============ ПАРСИНГ СУММЫ ============

def _extract_amount_from_question(text: str):
    cleaned = text
    for _ in range(6):
        new = re.sub(r'(\d)\s+(\d)', r'\1\2', cleaned)
        if new == cleaned:
            break
        cleaned = new

    amounts = re.findall(r'\d[\d,\.]*\d|\d{4,}', cleaned)
    results = []
    for a in amounts:
        a_clean = re.sub(r'[,\.](?=\d{3}(?:[,\.]|$))', '', a)
        a_clean = a_clean.replace(',', '.').replace(' ', '')
        try:
            val = float(a_clean)
            if val >= 100:
                results.append(val)
        except:
            pass

    return max(results) if results else None


# ============ CLAUDE AI ============

def ask_ai(question: str) -> str:
    if not ANTHROPIC_API_KEY:
        return "❌ ANTHROPIC_API_KEY не задан."

    today = datetime.now().strftime("%d.%m.%Y")

    # 1. Список названий счетов
    if _is_list_accounts_question(question):
        lines = ["Все счета компании:"]
        for i, acc in enumerate(MAIN_CASH_ACCOUNTS, 1):
            lines.append(f"{i}. {acc}")
        return "\n".join(lines)

    # 2. Конкретный счёт "Основная касса (Сейф)"
    if _is_seyf_question(question):
        try:
            spreadsheet = get_spreadsheet()
            реестр = spreadsheet.worksheet(SHEET_NAME)
            реестр_data = реестр.get_all_values()
            data_date = get_last_operation_date(реестр_data)
            initial_balances = _load_initial_balances()
            initial = _get_initial_for_account("Основная касса (Сейф)", initial_balances)
            ops_total = 0.0
            ops_count = 0
            for row in реестр_data[1:]:
                acc_val = str(row[6]).strip() if len(row) > 6 else ""
                amt_val = str(row[4]).strip() if len(row) > 4 else ""
                if not acc_val or not amt_val:
                    continue
                if _matches_account_strict(acc_val, "Основная касса (Сейф)"):
                    parsed = parse_amount_from_registry(amt_val)
                    if parsed is not None:
                        ops_total += parsed
                        ops_count += 1
            dds = round(initial + ops_total, 2)
            return (
                f"Основная касса (Сейф) на {data_date}: {dds:,.0f} тг\n"
                f"  Нач. остаток: {initial:,.0f} тг\n"
                f"  + Обороты ({ops_count} оп.): {ops_total:,.0f} тг"
            )
        except Exception as e:
            logger.error(f"seyf balance error: {e}")
            return f"❌ Ошибка при расчёте основной кассы (Сейф): {e}"

    # 3. Остаток на конкретный месяц
    if _is_balance_on_month_question(question):
        month_num = _extract_month_from_question(question)
        if month_num:
            try:
                now = datetime.now()
                year = now.year
                target_dt, label = _resolve_target_date_for_month(month_num, year)
                summary_text, _ = get_main_cash_summary_on_date(target_dt)
                if "последняя операция" in label:
                    summary_text = summary_text.replace(
                        f"Остатки по всем счетам на {target_dt.strftime('%d.%m.%Y')}:",
                        f"Остатки по всем счетам на {label}:"
                    )
                return summary_text
            except Exception as e:
                logger.error(f"balance_on_month error: {e}")
                return f"❌ Ошибка при расчёте остатка за месяц: {e}"

    # 4. Обороты/расходы/приходы за период
    _dates = _extract_dates_from_question(question)
    _has_turnover_intent = _is_turnover_intent_question(question)
    _has_money_with_date = _is_all_accounts_question(question) and len(_dates) > 0

    if _has_turnover_intent and not _dates:
        month_num = _extract_month_from_question(question)
        if month_num:
            now = datetime.now()
            year = now.year if month_num <= now.month else now.year - 1
            last_day = calendar.monthrange(year, month_num)[1]
            _dates = [f"01.{month_num:02d}.{year}", f"{last_day:02d}.{month_num:02d}.{year}"]

    if (_has_turnover_intent or _has_money_with_date) and len(_dates) > 0:
        try:
            if len(_dates) == 2:
                return get_turnover_for_range(_dates[0], _dates[1])
            else:
                t = question.lower()
                if any(kw in t for kw in [" на ", " к ", " до ", "к концу", "до конца"]):
                    year = datetime.now().year
                    start = f"01.01.{year}"
                    return get_turnover_for_range(start, _dates[0])
                else:
                    return get_turnover_for_range(_dates[0], _dates[0])
        except Exception as e:
            logger.error(f"get_turnover_for_range error: {e}")
            return f"❌ Ошибка при расчёте оборотов: {e}"

    # 5. Сводка по всем счетам
    if _is_all_accounts_question(question):
        month_filter = _extract_month_from_question(question)
        if month_filter:
            try:
                now = datetime.now()
                year = now.year
                target_dt, label = _resolve_target_date_for_month(month_filter, year)
                summary_text, _ = get_main_cash_summary_on_date(target_dt)
                return summary_text
            except Exception as e:
                logger.error(f"get_main_cash_summary_on_date error: {e}")
                return f"❌ Ошибка при расчёте за месяц: {e}"
        else:
            try:
                summary_text, _ = get_main_cash_summary()
                return summary_text
            except Exception as e:
                logger.error(f"get_main_cash_summary error: {e}")
                return f"❌ Ошибка при расчёте сводки: {e}"

    # 6. Поиск дня по остатку / обороту
    if _is_balance_search_question(question):
        target = _extract_amount_from_question(question)
        logger.info(f"Balance search: text='{question}' -> target={target}")
        if target is not None:
            try:
                matches = find_date_by_balance(target, tolerance=1.0)
                if matches:
                    lines = [f"Остаток {target:,.0f} тг найден:"]
                    for acc, date_str, bal in matches:
                        lines.append(f"  {acc}: {date_str} — {bal:,.0f} тг")
                    return "\n".join(lines)

                matches_daily = find_date_by_daily_total(target, tolerance=1.0)
                if matches_daily:
                    lines = [f"Найдено для суммы {target:,.0f} тг:"]
                    for acc, date_str, bal in matches_daily[:8]:
                        lines.append(f"  {acc}: {date_str} — {bal:,.0f} тг")
                    return "\n".join(lines)

                matches2 = find_date_by_balance(target, tolerance=5000.0)
                matches_daily2 = find_date_by_daily_total(target, tolerance=5000.0)
                all_approx = matches2 + matches_daily2
                if all_approx:
                    all_approx.sort(key=lambda x: abs(x[2] - target))
                    lines = [f"Точного совпадения нет. Ближайшие к {target:,.0f} тг:"]
                    for acc, date_str, bal in all_approx[:6]:
                        diff = int(round(bal - target))
                        sign = "+" if diff >= 0 else ""
                        lines.append(f"  {acc}: {date_str} — {bal:,.0f} тг ({sign}{diff:,})")
                    return "\n".join(lines)

                return f"Сумма {target:,.0f} тг не найдена ни в остатках, ни в оборотах реестра."
            except Exception as e:
                logger.error(f"Balance search error: {e}")
        else:
            logger.warning(f"_extract_amount_from_question вернул None для: '{question}'")

    tools = [
        {
            "name": "get_table_data",
            "description": (
                "Получить финансовые данные из Google Sheets реестра транзакций компании. "
                "Возвращает сводку (итоги по месяцам, счетам, статьям) и детальные строки. "
                "Используй для вопросов о финансах, остатках, оборотах, расходах, доходах. "
                "ВАЖНО: если вопрос касается конкретного месяца — обязательно передавай filter_month. "
                "Если вопрос касается конкретного счёта — передавай filter_account."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "filter_account": {
                        "type": "string",
                        "description": "Фильтр по названию счёта. Пусто = все счета."
                    },
                    "filter_month": {
                        "type": "string",
                        "description": (
                            "Номер месяца от 1 до 12. "
                            "Январь=1 ... Декабрь=12. Пусто = все месяцы. "
                            "ОБЯЗАТЕЛЬНО указывай если в вопросе упомянут конкретный месяц!"
                        )
                    },
                    "limit_rows": {
                        "type": "integer",
                        "description": "Количество детальных строк (по умолчанию 100, максимум 300).",
                        "default": 100
                    }
                },
                "required": []
            }
        },
        {
            "name": "get_balance_on_date",
            "description": (
                "Вычисляет остаток по конкретному счёту на указанную дату. "
                "Используй когда явно указан и счёт, и дата."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "account_name": {"type": "string", "description": "Название счёта"},
                    "date": {"type": "string", "description": "Дата в формате DD.MM.YYYY"}
                },
                "required": ["account_name", "date"]
            }
        },
        {
            "name": "get_all_accounts_summary",
            "description": (
                "Возвращает ТЕКУЩИЕ остатки и итог по всем счетам компании. "
                "ИСПОЛЬЗУЙ ДЛЯ: 'сколько денег', 'на всех счетах', "
                "'общий остаток', 'суммарный баланс', 'итого по всем счетам'. "
                "НЕ используй для вопросов про основную кассу (сейф). "
                "НЕ используй если в вопросе упомянут конкретный месяц — "
                "тогда используй get_all_accounts_summary_on_month."
            ),
            "input_schema": {"type": "object", "properties": {}, "required": []}
        },
        {
            "name": "get_all_accounts_summary_on_month",
            "description": (
                "Возвращает остатки по ВСЕМ счетам на конец указанного месяца. "
                "ИСПОЛЬЗУЙ когда спрашивают 'остаток на [месяц]', 'какой остаток на март', "
                "'сколько денег было в [месяц]', 'остатки на конец [месяц]', "
                "'сколько было на счетах в [месяц]', 'баланс на [месяц]'. "
                "Если месяц в прошлом — считает до последнего дня месяца. "
                "Если месяц текущий — считает до последней операции в реестре. "
                "Показывает ВСЕ счета без исключения. "
                "НЕ используй для вопросов про расходы/обороты — "
                "для них используй get_turnover_for_range."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "month": {
                        "type": "integer",
                        "description": "Номер месяца от 1 до 12. Январь=1 ... Декабрь=12."
                    },
                    "year": {
                        "type": "integer",
                        "description": "Год (например 2026). Если не указан — текущий год."
                    }
                },
                "required": ["month"]
            }
        },
        {
            "name": "get_seyf_balance",
            "description": (
                "Возвращает остаток по счёту Основная касса (Сейф). "
                "ИСПОЛЬЗУЙ ДЛЯ: 'основная касса', 'сейф', 'остаток кассы'. "
                "Это отдельный счёт наличных, НЕ сводка по всем счетам."
            ),
            "input_schema": {"type": "object", "properties": {}, "required": []}
        },
        {
            "name": "find_operation_by_amount",
            "description": (
                "Ищет операции в реестре по сумме. "
                "Используй когда спрашивают: 'какого дня эта сумма', 'когда была операция на X', "
                "'найди день с суммой X', 'какого числа было X тенге'. "
                "НЕ спрашивай уточнений — ищи сразу по всем счетам и всем датам."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "amount": {"type": "number", "description": "Сумма для поиска"},
                    "tolerance": {"type": "number", "description": "Допуск ±тг (по умолчанию 1.0)", "default": 1.0}
                },
                "required": ["amount"]
            }
        },
        {
            "name": "find_date_by_balance_tool",
            "description": (
                "Ищет дату когда остаток на счёте был равен указанной сумме. "
                "Используй когда спрашивают: 'какого дня остаток X', 'когда был остаток X тг', "
                "'какого числа на счёте было X'. "
                "Перебирает все счета и все даты автоматически."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "amount": {"type": "number", "description": "Искомый остаток в тенге"},
                    "tolerance": {"type": "number", "description": "Допуск ±тг (по умолчанию 1.0)", "default": 1.0}
                },
                "required": ["amount"]
            }
        },
        {
            "name": "get_turnover_for_range",
            "description": (
                "Считает приход и расход по всем счетам за указанный период. "
                "Для одного дня: start_date = end_date. "
                "ИСПОЛЬЗУЙ ДЛЯ: "
                "'сколько потрачено за 21 февраля' (start=end=21.02), "
                "'расход с 1 по 22 февраля' (start=01.02, end=22.02), "
                "'обороты за февраль' (start=01.02, end=28.02), "
                "'сколько потрачено в мае' (start=01.05, end=31.05). "
                "НЕ использовать для вопросов об остатках (балансах) — "
                "для остатков используй get_all_accounts_summary_on_month."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "start_date": {"type": "string", "description": "Начало периода DD.MM.YYYY"},
                    "end_date":   {"type": "string", "description": "Конец периода DD.MM.YYYY"}
                },
                "required": ["start_date", "end_date"]
            }
        },
        {
            "name": "get_daily_summary_for_period",
            "description": (
                "Строит ежедневную таблицу оборотов и остатков за указанный период. "
                "ИСПОЛЬЗУЙ ДЛЯ: 'покажи по дням за февраль', 'ежедневные остатки за январь', "
                "'динамика по дням', 'обороты по дням за период'. "
                "Возвращает строку на каждый день: приход, расход, нетто, итоговый остаток."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "start_date": {"type": "string", "description": "Начало периода DD.MM.YYYY"},
                    "end_date": {"type": "string", "description": "Конец периода DD.MM.YYYY"}
                },
                "required": ["start_date", "end_date"]
            }
        },
        {
            "name": "web_search",
            "description": "Поиск в интернете. Используй для погоды, курсов валют, новостей.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "query": {"type": "string", "description": "Поисковый запрос"}
                },
                "required": ["query"]
            }
        }
    ]

    system_prompt = (
        f"Ты финансовый помощник компании. Сегодня {today}.\n"
        "Правила:\n"
        "1. На финансовые вопросы используй get_table_data.\n"
        "2. КРИТИЧЕСКИ ВАЖНО: если вопрос про конкретный месяц — ВСЕГДА передавай filter_month.\n"
        "3. Если вопрос про конкретный счёт — передавай filter_account.\n"
        "4. На вопросы о погоде, курсах, новостях используй web_search.\n"
        "5. Отвечай ТОЛЬКО на русском языке. Будь кратким и точным.\n"
        "6. НЕ используй markdown: никаких **, #, _, списков с тире.\n"
        "7. Касса/Сейф — наличные компании, остаток может быть отрицательным — это нормально.\n"
        "8. Если спрашивают 'какого дня эта сумма X', 'когда было X тенге', "
        "   'какого числа операция на X' — СРАЗУ используй find_operation_by_amount. "
        "   НЕ спрашивай уточнений про счёт — ищи по всем счетам.\n"
        "9. Если спрашивают 'сколько денег', 'на всех счетах', 'общий остаток' БЕЗ указания дат — "
        "   используй get_all_accounts_summary.\n"
        "10. КРИТИЧЕСКИ ВАЖНО — 'остаток на [месяц]', 'какой остаток на [месяц]', "
        "    'баланс на [месяц]' — это get_all_accounts_summary_on_month. "
        "    Показывай ВСЕ счета, не только крупнейшие. "
        "    Если месяц в прошлом → последний день того месяца. "
        "    Если месяц текущий → дата последней операции в реестре.\n"
        "11. Если спрашивают 'обороты за [месяц]', 'приход/расход за [месяц]', 'статьи за [месяц]' — "
        "    используй get_table_data с filter_month.\n"
        "12. Если спрашивают 'основная касса', 'сейф', 'остаток кассы' — "
        "    используй get_seyf_balance. Это отдельный счёт, не сводка.\n"
        "13. Если указан конкретный счёт И дата — используй get_balance_on_date.\n"
        "14. Если спрашивают 'какого дня остаток X', 'когда был остаток X тг' — "
        "    используй find_date_by_balance_tool.\n"
        "15. В финансовых ответах дата расчёта берётся из данных (дата последней операции).\n"
        "16. ВАЖНО для вопросов про расходы/обороты с датами:\n"
        "    'потрачено ЗА 21 февраля' = get_turnover_for_range(start=21.02, end=21.02)\n"
        "    'потрачено С 1 ПО 22 февраля' = get_turnover_for_range(start=01.02, end=22.02)\n"
        "    'обороты за февраль' = get_turnover_for_range(start=01.02, end=28.02)\n"
        "17. Если спрашивают 'покажи по дням', 'ежедневные остатки', 'динамика по дням' — "
        "    используй get_daily_summary_for_period.\n"
        "18. КРИТИЧЕСКИ ВАЖНО — различай расходы и остатки:\n"
        "    'потрачено в мае', 'расход в мае' — get_turnover_for_range. НЕ остатки!\n"
        "    'остаток на май', 'баланс на май' — get_all_accounts_summary_on_month. НЕ обороты!\n"
        "19. ВСЕГДА показывай ВСЕ счета из списка при выводе остатков.\n"
    )

    messages = [{"role": "user", "content": question}]
    headers = {
        "x-api-key": ANTHROPIC_API_KEY,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }

    for _ in range(5):
        payload = {
            "model": "claude-sonnet-4-5",
            "max_tokens": 1500,
            "system": system_prompt,
            "tools": tools,
            "messages": messages,
        }
        try:
            resp = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers=headers, json=payload, timeout=45,
            )
            resp.raise_for_status()
            result = resp.json()
        except Exception as e:
            logger.error(f"Claude API error: {e}")
            return f"❌ Ошибка Claude API: {str(e)}"

        stop_reason = result.get("stop_reason")
        content = result.get("content", [])

        if stop_reason == "end_turn":
            text_parts = [b["text"] for b in content if b.get("type") == "text"]
            return "\n".join(text_parts).strip() or "Нет ответа."

        if stop_reason == "tool_use":
            messages.append({"role": "assistant", "content": content})
            tool_results = []

            for block in content:
                if block.get("type") != "tool_use":
                    continue
                tool_name = block["name"]
                tool_input = block.get("input", {})
                tool_use_id = block["id"]
                logger.info(f"Claude tool: {tool_name} {tool_input}")

                if tool_name == "get_table_data":
                    filter_account = tool_input.get("filter_account") or None
                    filter_month = tool_input.get("filter_month") or None
                    limit_rows = min(int(tool_input.get("limit_rows", 100)), 300)
                    summary, detail_rows = get_sheets_data_for_ai(
                        filter_account=filter_account,
                        filter_month=filter_month,
                        limit_rows=limit_rows,
                    )
                    if detail_rows:
                        detail_text = "\n\nДЕТАЛЬНЫЕ СТРОКИ:\n"
                        for r in detail_rows:
                            detail_text += (
                                f"  [{r['дата']}] {r['счёт']} | {r['сумма']} ₸"
                                f" | {r['статья']} | {r['описание']}\n"
                            )
                    else:
                        detail_text = "\n(По заданному фильтру строк не найдено)"
                    result_content = summary + detail_text

                elif tool_name == "get_balance_on_date":
                    acc = tool_input.get("account_name", "")
                    date_str = tool_input.get("date", "")
                    initial, ops_total, balance, ops_count = get_account_balance_on_date(acc, date_str)
                    if isinstance(ops_count, str):
                        result_content = f"Ошибка: {ops_count}"
                    else:
                        result_content = (
                            f"Остаток по счёту '{acc}' на {date_str}:\n"
                            f"  Нач. остаток: {initial:,.2f} тг\n"
                            f"  + Операции ({ops_count} строк): {ops_total:,.2f} тг\n"
                            f"  = Итого: {balance:,.2f} тг"
                        )

                elif tool_name == "get_all_accounts_summary":
                    summary_text, _ = get_main_cash_summary()
                    result_content = summary_text

                elif tool_name == "get_all_accounts_summary_on_month":
                    month = int(tool_input.get("month", datetime.now().month))
                    now = datetime.now()
                    year = int(tool_input.get("year", now.year))
                    try:
                        target_dt, label = _resolve_target_date_for_month(month, year)
                        summary_text, _ = get_main_cash_summary_on_date(target_dt)
                        if "последняя операция" in label:
                            summary_text = summary_text.replace(
                                f"Остатки по всем счетам на {target_dt.strftime('%d.%m.%Y')}:",
                                f"Остатки по всем счетам на {label}:"
                            )
                        result_content = summary_text
                    except Exception as e:
                        result_content = f"Ошибка расчёта остатков на конец месяца: {e}"

                elif tool_name == "get_seyf_balance":
                    spreadsheet = get_spreadsheet()
                    реестр = spreadsheet.worksheet(SHEET_NAME)
                    реестр_data = реестр.get_all_values()
                    data_date = get_last_operation_date(реестр_data)
                    initial_balances = _load_initial_balances()
                    initial = _get_initial_for_account("Основная касса (Сейф)", initial_balances)
                    ops_total = 0.0
                    ops_count = 0
                    for row in реестр_data[1:]:
                        acc_val = str(row[6]).strip() if len(row) > 6 else ""
                        amt_val = str(row[4]).strip() if len(row) > 4 else ""
                        if not acc_val or not amt_val:
                            continue
                        if _matches_account_strict(acc_val, "Основная касса (Сейф)"):
                            parsed = parse_amount_from_registry(amt_val)
                            if parsed is not None:
                                ops_total += parsed
                                ops_count += 1
                    dds = round(initial + ops_total, 2)
                    result_content = (
                        f"Основная касса (Сейф) на {data_date}: {dds:,.2f} тг\n"
                        f"  Нач. остаток: {initial:,.2f} тг\n"
                        f"  + Обороты ({ops_count} оп.): {ops_total:,.2f} тг"
                    )

                elif tool_name == "find_operation_by_amount":
                    amount = float(tool_input.get("amount", 0))
                    tolerance = float(tool_input.get("tolerance", 1.0))
                    matches = find_operation_by_amount(amount, tolerance=tolerance)
                    if matches:
                        lines = [f"Найдено операций с суммой {amount:,.0f} тг: {len(matches)}"]
                        for acc, date_str, amt, desc in matches[:10]:
                            sign = "+" if amt > 0 else ""
                            lines.append(f"  {date_str} | {acc} | {sign}{amt:,.0f} тг | {desc}")
                        if len(matches) > 10:
                            lines.append(f"  ... и ещё {len(matches) - 10} совпадений")
                        result_content = "\n".join(lines)
                    else:
                        matches2 = find_operation_by_amount(amount, tolerance=100.0)
                        if matches2:
                            lines = [f"Точного совпадения нет. Близкие к {amount:,.0f} тг:"]
                            for acc, date_str, amt, desc in matches2[:5]:
                                sign = "+" if amt > 0 else ""
                                lines.append(f"  {date_str} | {acc} | {sign}{amt:,.0f} тг | {desc}")
                            result_content = "\n".join(lines)
                        else:
                            result_content = f"Операций с суммой {amount:,.0f} тг не найдено."

                elif tool_name == "find_date_by_balance_tool":
                    amount = float(tool_input.get("amount", 0))
                    tolerance = float(tool_input.get("tolerance", 1.0))
                    matches = find_date_by_balance(amount, tolerance=tolerance)
                    matches_daily = find_date_by_daily_total(amount, tolerance=tolerance)
                    all_matches = matches + matches_daily
                    if all_matches:
                        lines = [f"Найдено для суммы {amount:,.0f} тг:"]
                        for acc, date_str, bal in all_matches[:10]:
                            lines.append(f"  {acc}: {date_str} — {bal:,.0f} тг")
                        result_content = "\n".join(lines)
                    else:
                        m2 = find_date_by_balance(amount, tolerance=5000.0)
                        m2d = find_date_by_daily_total(amount, tolerance=5000.0)
                        all_approx = m2 + m2d
                        if all_approx:
                            all_approx.sort(key=lambda x: abs(x[2] - amount))
                            lines = [f"Точного совпадения нет. Ближайшие к {amount:,.0f} тг:"]
                            for acc, date_str, bal in all_approx[:6]:
                                diff = int(round(bal - amount))
                                sign = "+" if diff >= 0 else ""
                                lines.append(f"  {acc}: {date_str} — {bal:,.0f} тг ({sign}{diff:,})")
                            result_content = "\n".join(lines)
                        else:
                            result_content = f"Сумма {amount:,.0f} тг не найдена."

                elif tool_name == "get_turnover_for_range":
                    result_content = get_turnover_for_range(
                        tool_input.get("start_date", ""),
                        tool_input.get("end_date", "")
                    )

                elif tool_name == "get_daily_summary_for_period":
                    result_content = get_daily_summary_for_period(
                        tool_input.get("start_date", ""),
                        tool_input.get("end_date", "")
                    )

                elif tool_name == "web_search":
                    result_content = _do_web_search(tool_input.get("query", ""))

                else:
                    result_content = f"Инструмент '{tool_name}' не найден."

                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": tool_use_id,
                    "content": result_content,
                })

            messages.append({"role": "user", "content": tool_results})
        else:
            text_parts = [b["text"] for b in content if b.get("type") == "text"]
            return "\n".join(text_parts).strip() or "Нет ответа."

    return "❌ Превышено число итераций. Попробуйте переформулировать вопрос."


def _do_web_search(query: str) -> str:
    try:
        url = "https://api.duckduckgo.com/"
        params = {"q": query, "format": "json", "no_html": "1", "skip_disambig": "1"}
        resp = requests.get(url, params=params, timeout=10)
        data = resp.json()
        results = []
        abstract = data.get("AbstractText", "").strip()
        if abstract:
            results.append(f"Краткий ответ: {abstract}")
        answer = data.get("Answer", "").strip()
        if answer:
            results.append(f"Ответ: {answer}")
        for item in data.get("RelatedTopics", [])[:3]:
            if isinstance(item, dict) and item.get("Text"):
                results.append(item["Text"])
        return "\n".join(results) if results else f"По запросу '{query}' ничего не найдено."
    except Exception as e:
        return f"Ошибка поиска: {str(e)}"


# ============ PDF Kaspi Gold ============

def _parse_kaspi_text_lines(all_text, account):
    rows = []
    lines = all_text.split("\n")
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        m = re.match(
            r'^(\d{2}\.\d{2}\.\d{2,4})\s+([+\-])\s*([\d\s]+,\d{2})\s*₸\s+(.+)$',
            line
        )
        if m:
            date_raw = m.group(1)
            sign = 1 if m.group(2) == '+' else -1
            amt_str = m.group(3).replace(" ", "").replace(",", ".")
            desc = m.group(4).strip()
            j = i + 1
            while j < len(lines):
                nxt = lines[j].strip()
                if not nxt:
                    j += 1
                    continue
                if re.match(r'^\d{2}\.\d{2}\.\d{2,4}', nxt):
                    break
                if re.match(r'^\([-+]?\s*[\d\s]+[,.][\d]+\s*(USD|EUR|RUB|CNY)\)', nxt, re.IGNORECASE):
                    desc = f"{desc} {nxt}".strip()
                    j += 1
                    continue
                if nxt.startswith("("):
                    j += 1
                    continue
                desc = f"{desc} {nxt}".strip()
                j += 1
                break
            i = j
            try:
                amount = sign * float(amt_str)
                date_str = format_date(date_raw)
                rows.append(make_row(date_str, amount, account, desc))
            except:
                pass
        else:
            i += 1
    return rows


def process_kaspi_gold_pdf(file_bytes):
    rows = []
    account = "Каспи Голд Айко"
    closing_balance = None
    opening_balance = None

    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        first_text = pdf.pages[0].extract_text() or ""
        all_text = ""
        for p in pdf.pages:
            all_text += (p.extract_text() or "") + "\n"

        is_deposit = "По Депозиту" in first_text or "На Депозите" in first_text

        m_iban = re.search(r"Номер счета[:\s]+(KZ\w+)", all_text)
        if not m_iban:
            m_iban = re.search(r"счет[оом]*\s+(KZ\w+)", all_text, re.IGNORECASE)
        if m_iban:
            iban = m_iban.group(1).strip()
            account = IBAN_MAP.get(iban, account)

        if not is_deposit:
            bal_matches = re.findall(
                r"Доступно на\s+(\d{2}\.\d{2}\.\d{2,4})[:\s]*[+\-]?\s*([\d\s]+,\d+)\s*₸",
                all_text
            )
            if len(bal_matches) >= 2:
                opening_balance = parse_num(bal_matches[0][1])
                closing_balance = parse_num(bal_matches[-1][1])
            elif len(bal_matches) == 1:
                closing_balance = parse_num(bal_matches[0][1])

            rows = _parse_kaspi_text_lines(all_text, account)

            if not rows:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    for table in tables:
                        for row in table:
                            if not row or len(row) < 3:
                                continue
                            date_cell = str(row[0] or "").strip()
                            amount_cell = str(row[1] or "").strip()
                            op_cell = str(row[2] or "").strip() if len(row) > 2 else ""
                            det_cell = str(row[3] or "").strip() if len(row) > 3 else ""
                            desc_cell = f"{op_cell} {det_cell}".strip()
                            if not re.match(r"\d{2}\.\d{2}\.\d{2,4}", date_cell):
                                continue
                            amount_clean = (amount_cell.replace(" ", "").replace("₸", "")
                                            .replace("\xa0", "").replace(",", "."))
                            sign = 1
                            if amount_clean.startswith("-"):
                                sign = -1
                                amount_clean = amount_clean[1:]
                            elif amount_clean.startswith("+"):
                                amount_clean = amount_clean[1:]
                            try:
                                amount = sign * float(amount_clean)
                            except:
                                continue
                            rows.append(make_row(format_date(date_cell), amount, account, desc_cell))
        else:
            dep_matches = re.findall(
                r"На Депозите\s+\d{2}\.\d{2}\.\d{2,4}\s+([\d\s]+[,.][\d]+)\s*₸",
                all_text
            )
            if len(dep_matches) >= 2:
                opening_balance = parse_num(dep_matches[0])
                closing_balance = parse_num(dep_matches[-1])
            elif len(dep_matches) == 1:
                closing_balance = parse_num(dep_matches[0])
            rows = _parse_kaspi_text_lines(all_text, account)

    if opening_balance is None and closing_balance is not None and rows:
        total_ops = sum(float(r[4]) for r in rows)
        opening_balance = round(closing_balance - total_ops, 2)

    logger.info(f"Kaspi Gold: счет={account}, строк={len(rows)}, вх={opening_balance}, исх={closing_balance}")
    return rows, account, closing_balance, opening_balance


# ============ PDF BCC ============

def process_bcc_pdf(file_bytes):
    rows = []
    account = "БЦК Камкорлык"
    closing_balance = None
    opening_balance = None

    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        full_text = ""
        for page in pdf.pages:
            full_text += (page.extract_text() or "") + "\n"

        m = re.search(r"ЖСК\s*/\s*ИИК\s*:\s*(KZ\w+)", full_text)
        if not m:
            m = re.search(r"ЖСК\s*/\s*ИИК\s+(KZ\w+)", full_text)
        if not m:
            m = re.search(r"IBAN[:\s]+(KZ\w+)", full_text)
        if m:
            account = IBAN_MAP.get(m.group(1).strip(), m.group(1).strip())

        m_open = re.search(
            r"[Кк]іріс [қк]алдық\s*/\s*[Вв]ходящий остаток[:\s]*([\d\s]+[,.][\d]+)", full_text)
        if not m_open:
            m_open = re.search(
                r"[Кк]іріс сальдо\s*/\s*[Вв]ходящее сальдо[:\s]*([\d\s]+[,.][\d]+)", full_text)
        if m_open:
            opening_balance = parse_num(m_open.group(1))

        m_bal = re.search(
            r"[Шш]ығыс сальдо\s*/\s*[Ии]сходящее сальдо[:\s]*([\d\s]+[,.][\d]+)", full_text)
        if not m_bal:
            m_bal = re.search(r"[Ии]сходящее сальдо[:\s]*([\d\s]+[,.][\d]+)", full_text)
        if not m_bal:
            m_bal = re.search(r"[Шш]ығыс сальдо[:\s]*([\d\s]+[,.][\d]+)", full_text)
        if m_bal:
            closing_balance = parse_num(m_bal.group(1))

        all_table_rows = []
        for page in pdf.pages:
            for table in page.extract_tables():
                for row in table:
                    if row and len(row) >= 12:
                        all_table_rows.append(list(row))

        merged_rows = []
        i = 0
        while i < len(all_table_rows):
            row = all_table_rows[i]
            c0 = str(row[0] or "").replace("\n", "").strip()
            c1 = str(row[1] or "").replace("\n", "").strip()
            if "итого" in c0.lower() or "жиынтығы" in c0.lower():
                i += 1
                continue
            if "дата" in c1.lower() or "күні" in c1.lower() or "реттік" in c0.lower():
                i += 1
                continue
            if c0 == "NT-" and i + 1 < len(all_table_rows):
                next_row = all_table_rows[i + 1]
                nc0 = str(next_row[0] or "").replace("\n", "").strip()
                if re.match(r"^\d+$", nc0):
                    glued = []
                    for ci in range(12):
                        pv = str(row[ci] or "").replace("\n", " ").strip()
                        nv = str(next_row[ci] or "").replace("\n", " ").strip()
                        if ci == 0:
                            glued.append("NT-" + nc0)
                        elif ci == 1:
                            if re.search(r"\d{1,2}\.\d{2}\.\d{4}", pv):
                                glued.append(pv)
                            elif re.search(r"\d{1,2}\.\d{2}\.\d{4}", nv):
                                glued.append(nv)
                            else:
                                glued.append((pv + nv).strip())
                        elif ci in (7, 8):
                            glued.append((pv + nv).replace(" ", ""))
                        else:
                            glued.append((pv + " " + nv).strip() if (pv and nv) else (pv or nv))
                    merged_rows.append(glued)
                    i += 2
                    continue
            merged_rows.append(row)
            i += 1

        for row in merged_rows:
            c0 = str(row[0] or "").replace("\n", "").strip()
            c1 = str(row[1] or "").replace("\n", "").strip()
            c7 = str(row[7] or "").replace("\n", " ").strip()
            c8 = str(row[8] or "").replace("\n", " ").strip()
            c11 = str(row[11] or "").replace("\n", " ").strip()
            c7 = re.sub(r'\(.*?(?:USD|EUR|RUB|CNY).*?\)', '', c7, flags=re.IGNORECASE).strip()
            c8 = re.sub(r'\(.*?(?:USD|EUR|RUB|CNY).*?\)', '', c8, flags=re.IGNORECASE).strip()
            if "итого" in c1.lower() or "жиынтығы" in c1.lower():
                continue
            date_m = re.search(r"(\d{1,2})\.(\d{2})\.(\d{4})", c1)
            if not date_m:
                continue
            day, month, year = int(date_m.group(1)), int(date_m.group(2)), date_m.group(3)
            date_str = f"{month:02d}/{day:02d}/{year}"
            debit = parse_num(c7)
            credit = parse_num(c8)
            if debit > 0:
                amount = -debit
            elif credit > 0:
                amount = credit
            else:
                continue
            desc = f"{c0} {c11}".strip() if c0 else c11
            rows.append(make_row(date_str, amount, account, desc))

    logger.info(f"BCC: счет={account}, строк={len(rows)}, вх={opening_balance}, исх={closing_balance}")
    return rows, account, closing_balance, opening_balance


# ============ PDF Halyk ============

def parse_kz_num(s):
    s = str(s or "").strip().replace(" ", "").replace("\xa0", "")
    s = re.sub(r",(\d{3})(?=[\d,.])", r"\1", s)
    s = s.replace(",", ".")
    try:
        return float(s)
    except:
        return 0


def process_halyk_pdf(file_bytes):
    rows = []
    account = "Народный банк"
    closing_balance = None
    opening_balance = None

    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        full_text = ""
        for page in pdf.pages:
            full_text += (page.extract_text() or "") + "\n"

    m_iban = re.search(r"Счет\(Валюта\)[:\s]+(KZ[\w]+)", full_text)
    if m_iban:
        account = IBAN_MAP.get(m_iban.group(1).strip(), account)

    m_open = re.search(r"[Вв]ходящий остаток[:\s]*([\d\s,]+\.\d{2})", full_text)
    if m_open:
        opening_balance = parse_kz_num(m_open.group(1))

    m_bal = re.search(r"[Ии]сходящий остаток[:\s]*([\d\s,]+\.\d{2})", full_text)
    if m_bal:
        closing_balance = parse_kz_num(m_bal.group(1))

    lines = full_text.split("\n")
    blocks = []
    current = []
    for line in lines:
        line = line.strip()
        if re.match(r"^\d{2}\.\d{2}\.\d{4}\s+\S+\s+[\d,]+\.\d{2}", line):
            if current:
                blocks.append(current)
            current = [line]
        elif current:
            current.append(line)
    if current:
        blocks.append(current)

    for block in blocks:
        first = block[0]
        full_desc = " ".join(block)
        m = re.match(r"^(\d{2}\.\d{2}\.\d{4})\s+\S+\s+([\d,]+\.\d{2})", first)
        if not m:
            continue
        date_raw = m.group(1)
        amount = parse_kz_num(m.group(2))
        if any(kw in full_desc.lower() for kw in ["снятие", "комиссия", "перевод", "cmstake"]):
            amount = -amount
        d, mo, y = date_raw.split(".")
        date_str = f"{int(mo):02d}/{int(d):02d}/{y}"
        doc_num_m = re.match(r'^\d{2}\.\d{2}\.\d{4}\s+(\S+)\s+', first)
        doc_num = doc_num_m.group(1) if doc_num_m else ""
        ubs_m = re.search(r'00UBS(\d+)', full_desc)
        ubs_num = ("00UBS" + ubs_m.group(1)) if ubs_m else ""
        clean_desc = re.sub(r'^\d{2}\.\d{2}\.\d{4}\s+\S+\s+[\d,]+\.\d{2}\s*', '', full_desc).strip()
        if ubs_num:
            desc_final = f"{ubs_num} {doc_num} {clean_desc}".strip()
        elif doc_num:
            desc_final = f"{doc_num} {clean_desc}".strip()
        else:
            desc_final = clean_desc
        rows.append(make_row(date_str, amount, account, desc_final[:200]))

    return rows, account, closing_balance, opening_balance


# ============ XLSX ============

def process_xlsx(file_bytes):
    rows = []
    closing_balance = None
    opening_balance = None

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    iban = str(cell_val(ws.cell(row=3, column=3))).strip()
    account = IBAN_MAP.get(iban, iban)

    opening_val = cell_val(ws.cell(row=9, column=3))
    try:
        opening_balance = float(str(opening_val).replace(" ", "").replace(",", "."))
    except:
        opening_balance = None

    closing_val = cell_val(ws.cell(row=10, column=3))
    try:
        closing_balance = float(str(closing_val).replace(" ", "").replace(",", "."))
    except:
        pass

    if not closing_balance:
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = str(cell_val(ws.cell(row=row_idx, column=col_idx))).lower()
                if "исходящ" in cell and "сальдо" in cell:
                    for c in range(col_idx + 1, min(col_idx + 5, ws.max_column + 1)):
                        v = cell_val(ws.cell(row=row_idx, column=c))
                        if v:
                            try:
                                closing_balance = float(str(v).replace(" ", "").replace(",", "."))
                                break
                            except:
                                pass
                    if closing_balance:
                        break
            if closing_balance:
                break

    data_start = 14
    for row_idx in range(12, min(20, ws.max_row + 1)):
        val = cell_val(ws.cell(row=row_idx, column=2))
        if val and re.search(r"\d{1,2}[.\-/]\d{1,2}[.\-/]\d{2,4}", str(val)):
            data_start = row_idx
            break

    for row_idx in range(data_start, ws.max_row + 1):
        date_val = cell_val(ws.cell(row=row_idx, column=2))
        debit = cell_val(ws.cell(row=row_idx, column=3))
        credit = cell_val(ws.cell(row=row_idx, column=4))
        supplier = str(cell_val(ws.cell(row=row_idx, column=5)) or "")
        desc_raw = str(cell_val(ws.cell(row=row_idx, column=9)) or "")
        doc_num_val = str(cell_val(ws.cell(row=row_idx, column=1)) or "").strip()
        if doc_num_val and re.match(r'^\d{7,}$', doc_num_val):
            desc = f"{doc_num_val} {desc_raw}".strip()
        else:
            desc = desc_raw
        if not date_val:
            continue
        date_str = format_date(date_val)
        if not re.search(r"\d{2}/\d{2}/\d{4}", date_str):
            continue
        has_d = debit not in ("", None, "None")
        has_c = credit not in ("", None, "None")
        if has_d:
            try:
                amount = -float(str(debit).replace(" ", "").replace(",", "."))
            except:
                continue
        elif has_c:
            try:
                amount = float(str(credit).replace(" ", "").replace(",", "."))
            except:
                continue
        else:
            continue
        rows.append(make_row(date_str, amount, account, desc, supplier))

    return rows, account, closing_balance, opening_balance


# ============ ОПРЕДЕЛЕНИЕ ТИПА PDF ============

def detect_pdf_type(first_page_text):
    if "По Депозиту" in first_page_text or "На Депозите" in first_page_text:
        return "kaspi_deposit"
    if ("СПРАВКА" in first_page_text
            and ("Kaspi Gold" in first_page_text or "CASPKZKA" in first_page_text)):
        return "kaspi_gold"
    if "Kaspi Gold" in first_page_text:
        return "kaspi_gold"
    if "Kaspi Bank" in first_page_text or "CASPKZKA" in first_page_text:
        return "kaspi_gold"
    if ("Народный Банк" in first_page_text or "Halyk" in first_page_text
            or "HSBKKZKX" in first_page_text):
        return "halyk"
    if ("ЦентрКредит" in first_page_text or "ЦентрКре" in first_page_text
            or "KCJBKZKX" in first_page_text
            or "KZ25856" in first_page_text):
        return "bcc"
    return "kaspi_gold"


# ============ TELEGRAM HANDLER ============

async def handle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message:
        return

    if update.message.text:
        question = update.message.text.strip()

        if question.startswith("/start"):
            await update.message.reply_text(
                "👋 Привет! Я бухгалтерский бот.\n\n"
                "📎 Отправьте файл выписки (.xlsx или .pdf) — загружу в таблицу.\n\n"
                "💬 Примеры вопросов:\n"
                "• Сколько денег на всех счетах?\n"
                "• Остаток на март / Какой остаток на май?\n"
                "• Сколько денег было в январе?\n"
                "• Сколько потрачено в мае?\n"
                "• Обороты за март по статьям?\n"
                "• Основная касса (Сейф) — сколько?\n"
                "• Покажи обороты по дням за январь\n"
                "• Какого дня был остаток 685486 тг?\n"
                "• Какой курс доллара?\n"
                "• Напиши все счета\n\n"
                "🔍 Команды:\n"
                "/rows <счёт> — строки по счёту\n"
                "Пример: /rows Каспи Айс"
            )
            return

        if question.startswith("/rows"):
            acc = question[5:].strip()
            if not acc:
                await update.message.reply_text("Укажи счёт: /rows Каспи Айс")
                return
            await update.message.reply_text(f"🔍 Ищу строки для «{acc}»...")
            try:
                matched = get_account_rows(acc)
            except Exception as e:
                await update.message.reply_text(f"❌ Ошибка: {e}")
                return
            if not matched:
                await update.message.reply_text(f"❌ Строк по счёту «{acc}» не найдено.")
                return
            total = 0.0
            реестр_data_for_date = get_spreadsheet().worksheet(SHEET_NAME).get_all_values()
            data_date = get_last_operation_date(реестр_data_for_date)
            for _, row in matched:
                try:
                    val = parse_amount_from_registry(row[4])
                    if val is not None:
                        total += val
                except:
                    pass
            header = (
                f"📋 Счёт: {acc}\n"
                f"На дату: {data_date}\n"
                f"Найдено строк: {len(matched)}\n"
                f"Сумма операций: {total:,.2f} ₸\n"
                f"{'─' * 35}\n"
            )
            chunk_size = 50
            chunks = [matched[i:i + chunk_size] for i in range(0, len(matched), chunk_size)]
            for part_idx, chunk in enumerate(chunks):
                lines = []
                for sheet_row, row in chunk:
                    date = row[3] if len(row) > 3 else ""
                    amount = row[4] if len(row) > 4 else ""
                    desc = row[8] if len(row) > 8 else ""
                    desc_short = (desc[:40] + "…") if len(desc) > 40 else desc
                    lines.append(f"#{sheet_row} | {date} | {amount} | {desc_short}")
                part_text = "\n".join(lines)
                msg = (header + part_text) if part_idx == 0 else f"(продолжение {part_idx + 1}/{len(chunks)})\n{part_text}"
                await update.message.reply_text(msg)
            return

        await update.message.reply_text("🤔 Думаю...")
        try:
            answer = ask_ai(question)
        except Exception as e:
            logger.error(f"ask_ai error: {e}")
            answer = f"❌ Ошибка ИИ: {str(e)}"
        await update.message.reply_text(answer)
        return

    doc = update.message.document
    if not doc:
        return

    fname = (doc.file_name or "").lower()
    if not (fname.endswith(".xlsx") or fname.endswith(".pdf")):
        await update.message.reply_text("Поддерживаются только .xlsx и .pdf файлы")
        return

    await update.message.reply_text("⏳ Обрабатываю выписку...")

    try:
        file = await context.bot.get_file(doc.file_id)
        file_bytes = bytes(await file.download_as_bytearray())

        if fname.endswith(".xlsx"):
            rows, account, closing_balance, opening_balance = process_xlsx(file_bytes)
        else:
            with pdfplumber.open(BytesIO(file_bytes)) as pdf:
                first_page_text = pdf.pages[0].extract_text() or ""
            pdf_type = detect_pdf_type(first_page_text)
            logger.info(f"PDF тип: {pdf_type}")
            if pdf_type in ("kaspi_gold", "kaspi_deposit"):
                rows, account, closing_balance, opening_balance = process_kaspi_gold_pdf(file_bytes)
            elif pdf_type == "halyk":
                rows, account, closing_balance, opening_balance = process_halyk_pdf(file_bytes)
            elif pdf_type == "bcc":
                rows, account, closing_balance, opening_balance = process_bcc_pdf(file_bytes)
            else:
                rows, account, closing_balance, opening_balance = process_kaspi_gold_pdf(file_bytes)

        if not rows:
            await update.message.reply_text("❌ Операции не найдены в файле")
            return

        time.sleep(2)
        sheet = get_sheet()
        existing_data = sheet.get_all_values()
        key_to_row, existing_keys = build_existing_keys(existing_data)

        dupe_sheet_rows = []
        dupe_rows = []
        for r in rows:
            if is_duplicate(r, existing_keys):
                dupe_rows.append(r)
                dupe_sheet_rows.append(find_existing_row(r, key_to_row))
        dupes = len(dupe_rows)

        def format_row_ranges(row_nums):
            if not row_nums:
                return ""
            sorted_nums = sorted(n for n in row_nums if n is not None)
            if not sorted_nums:
                return ""
            ranges = []
            start = end = sorted_nums[0]
            for n in sorted_nums[1:]:
                if n == end + 1:
                    end = n
                else:
                    ranges.append(f"{start}-{end}" if start != end else str(start))
                    start = end = n
            ranges.append(f"{start}-{end}" if start != end else str(start))
            return ", ".join(ranges)

        all_vals_for_count = sheet.get_all_values()
        last_filled = 1
        for i, row in enumerate(all_vals_for_count, start=1):
            if any(str(c).strip() for c in row):
                last_filled = i
        next_row = last_filled + 1

        if dupes == len(rows):
            dupe_range = format_row_ranges(dupe_sheet_rows)
            msg = (
                f"⚠️ Этот файл уже был загружен ранее!\n"
                f"Все {len(rows)} строк уже есть в таблице.\n"
                f"Строки в таблице: {dupe_range}\nНичего не добавлено."
            )
            msg += build_balance_msg(account, closing_balance)
            msg += f"\n\n🔗 {SPREADSHEET_URL}"
            await update.message.reply_text(msg)
            return

        if dupes > 0:
            dupe_range = format_row_ranges(dupe_sheet_rows)
            new_rows = [r for r in rows if not is_duplicate(r, existing_keys)]
            new_range = format_row_ranges(list(range(next_row, next_row + len(new_rows))))
            rows = new_rows
            await update.message.reply_text(
                f"⚠️ {dupes} строк уже есть в таблице\n"
                f"Дубли в строках: {dupe_range}\n"
                f"Добавляю {len(rows)} новых → строки {new_range}"
            )

        rows.sort(key=lambda r: datetime.strptime(r[3], "%m/%d/%Y") if r[3] else datetime.min)
        actual_start = append_rows_from_col_a(sheet, rows)
        time.sleep(3)

        added_range = format_row_ranges(list(range(actual_start, actual_start + len(rows))))
        msg = f"✅ Готово! Добавлено {len(rows)} строк\nСчет: {account}\n📋 Строки: {added_range}\n"
        msg += build_balance_msg(account, closing_balance)
        msg += f"\n\n🔗 {SPREADSHEET_URL}"
        await update.message.reply_text(msg)

    except Exception as e:
        logger.error(f"Error: {e}")
        await update.message.reply_text(f"❌ Ошибка: {str(e)}")


def main():
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(MessageHandler(filters.ALL, handle))
    logger.info("Bot started!")
    app.run_webhook(
        listen="0.0.0.0",
        port=PORT,
        webhook_url=f"{WEBHOOK_URL}/webhook",
        url_path="webhook",
        drop_pending_updates=True,
    )


if __name__ == "__main__":
    main()
